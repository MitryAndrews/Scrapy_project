import os
import csv
import shutil
import zipfile
import openpyxl
from openpyxl import load_workbook
move_dir = fr'C:\Users\update_xlsx_files' # folder where we move files
dirname_from = fr'C:\Users\update'
dirname = fr'C:\Users\update_xlsx_files' #
# dirname = fr'C:\Users\update'
files = os.listdir(dirname_from)
print((files))
temp = map(lambda name: os.path.join(dirname_from, name), files)
keys =['99. Файл', '99. Реестр платежей №', '99. Номер расч.счёта', '99. Наименование Застройщика', '99. Дата реестра', '99. по объекту строительства']
dict_value = {}
dict_list = []
other_dict = []
print(list(temp), sep='\n')
# print(os.getcwd())
tree = list(os.walk(dirname_from))
list_file_xlsx = []
for i in tree:
    print(i)
for adress, dirs, files in os.walk(dirname_from):
    for name in files:
        extention = os.path.splitext(name)[-1].replace('.', '')
        print(adress)
        if extention == 'xlsx':
            # path_file = os.path.join(adress, name)
            # list_file_xlsx.append(path_file)

            path_file = os.path.join(adress, name)
            list_file_xlsx.append(path_file)
            dst = os.path.join(move_dir, name)
            print('path== ', path_file, '   ==    new_file ==', dst)
            shutil.copy(path_file, dst)

            print('extention== ', extention)
print(len(list_file_xlsx), list_file_xlsx)
'==========================================================='
count = 0
for name_file in list_file_xlsx:
    # print(name_file)
    wb = load_workbook(name_file)
    # print(wb.sheetnames)
    # print(wb.active)
    ws = wb.active
    total = []
    count += 1
    # print(f'file===> {count} ===> ', name_file, ws['A2'].value, ws['A4'].value, ws['A6'].value, ws['A7'].value, sep='\n')
    for n in range(1, ws.max_row):
        d = ws.cell(row=n, column=7)
        ddd = str(d.value)
        if 'SUM' in ddd:
            ddd = 'ИТОГО'
            cell_max_row = d.row
        total.append(ddd)
        # print(ddd, type(ddd))
    dd = list(set(total))
    if 'ИТОГО' not in dd:
        print(dd)
        print(f'file===> {count} ===> ', name_file)
    for i in range(1, ws.max_row):
        for col in ws.iter_cols(1, ws.max_column):
            adress_cell = str(col[i]).replace('<', '').replace('>', '').split('.')[-1]
            cell_value = col[i].value
            # print(cell_value, end='\t\t')
            if cell_value == '№ п/п':
                cell_min_row = col[i].row
                # print(col[i].value, col[i].column)
            if cell_value == 'ИТОГО':
                cell_max_row = col[i].row
            if cell_value == 'Статья расчета общей стоимости строительного проекта*':
                cell_max_col = col[i].column
    for i in range(cell_min_row + 1, cell_max_row - 1):
        count_keys = 1
        dict_value = {}
        for col in ws.iter_cols(1, cell_max_col):
            keys_n = ws.cell(row=cell_min_row, column=count_keys)
            keys_n = keys_n.value
            count_keys2 = '0' + str(count_keys) if len(str(count_keys)) == 1 else count_keys
            if keys_n == None:
                keys_reg = (str(count_keys2)+ '. ' + 'None')
            else:
                keys_reg = (str(count_keys2)+ '. ' + keys_n)
            keys.append(keys_reg)
            count_keys += 1
            # print(i)
            x = str(col[i].value).replace('\n', '')
            dict_value['99. Файл'] = name_file
            dict_value['99. Реестр платежей №'] = ws['A2'].value.replace('\n', '').replace('РЕЕСТР ПЛАТЕЖЕЙ','').strip()
            dict_value['99. Номер расч.счёта'] = str(ws['A4'].value.replace('Номер расчетного счета Застройщика, с которого производятся платежи:', '').replace('\n', '').strip())
            dict_value['99. Наименование Застройщика'] = str(ws['A5'].value.replace('Наименование Застройщика','').replace('\n', '').strip())
            dict_value['99. Дата реестра'] = str(ws['A6'].value.replace('«', '').replace('»', '').replace('г.', '').replace('\n', '').strip())
            dict_value['99. по объекту строительства'] = ws['A7'].value.replace('по объекту строительства:','').replace('\n', '').strip()
            dict_value[keys_reg] = x
        dict_list.append(dict_value)
        print(f'file===> {count} ===> ', name_file)
        # print('==== >', dict_value)
        # print(dict_list)
        # print(dict_list)
        # dict_list.append(dict_value)
        # print(dict_list)
        #     print(col[i].value, end='\t\t')
        # print('')
keys = list(set(keys))
keys.sort()
# print(dict_list)

with open(fr'C:\Users\user\Desktop\Рабочие файлы\Registry_base_add.csv', 'w', encoding='utf-8', newline='') as csv_file:
    fieldnames = keys
    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(dict_list)
    # for i in dict_list:
    #     # print(i)
    #     writer.writerow(i)

# print(cell_min_row, ' -- ', cell_max_row, ' -- ', cell_max_col)


# print('=========================================')
# for i in range(cell_min_row + 1, cell_max_row-1):
#     for col in ws.iter_cols(1, cell_max_col):
#         print(col[i].value, end='\t\t')
#     print('')

